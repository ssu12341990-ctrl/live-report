from __future__ import annotations

from datetime import datetime, time
from typing import Any

import openpyxl


ADQ_INPUT_FIELDS = [
    "日期",
    "主播",
    "开播时间",
    "结束时间",
    "直播时长(h)",
    "场次ID",
    "投流组名",
    "投流渠道",
    "创意名称",
    "主要地域",
    "观看人数",
    "花费(元)",
    "成交量",
    "转化目标成本(元)",
    "商品点击",
    "提交订单",
    "支付成功",
    "互动人数",
    "关注人数",
    "支付金额(元)",
    "退款金额(元)",
    "备注",
]

OPTIONAL_FIELDS = {"关注人数", "支付金额(元)", "退款金额(元)", "备注"}
NUMERIC_FIELDS = {
    "直播时长(h)",
    "观看人数",
    "花费(元)",
    "成交量",
    "转化目标成本(元)",
    "商品点击",
    "提交订单",
    "支付成功",
    "互动人数",
    "关注人数",
    "支付金额(元)",
    "退款金额(元)",
}


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d %H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def _as_number(value: Any, *, required: bool, field: str, row_idx: int) -> float:
    if value in (None, ""):
        if required:
            raise ValueError(f"第 {row_idx} 行缺少必填数值字段：{field}")
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "":
        if required:
            raise ValueError(f"第 {row_idx} 行缺少必填数值字段：{field}")
        return 0.0
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"第 {row_idx} 行字段 {field} 不是有效数值：{value}") from exc


def _time_to_hour(value: Any) -> float | None:
    if isinstance(value, datetime):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    text = _as_text(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.hour + parsed.minute / 60 + parsed.second / 3600
        except ValueError:
            continue
    return None


def _resolve_duration_hours(row: dict[str, Any], row_idx: int) -> float:
    duration = row.get("直播时长(h)")
    if duration not in (None, ""):
        return _as_number(duration, required=False, field="直播时长(h)", row_idx=row_idx)

    start_h = _time_to_hour(row.get("开播时间"))
    end_h = _time_to_hour(row.get("结束时间"))
    if start_h is None or end_h is None:
        return 0.0
    if end_h < start_h:
        end_h += 24.0
    return round(end_h - start_h, 2)


def _select_input_sheet(workbook: openpyxl.Workbook):
    preferred_names = ("ADQ输入模板", "adq_input", "Sheet1")
    for name in preferred_names:
        if name in workbook.sheetnames:
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def load_adq_rows(path: str) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = _select_input_sheet(workbook)

    header_row = [cell.value for cell in sheet[1]]
    header_map = {_as_text(value): idx for idx, value in enumerate(header_row)}

    missing = [field for field in ADQ_INPUT_FIELDS if field not in header_map]
    if missing:
        raise ValueError(f"输入模板缺少字段：{', '.join(missing)}")

    data_rows: list[dict[str, Any]] = []
    for row_idx in range(5, sheet.max_row + 1):
        row_values = [sheet.cell(row=row_idx, column=idx + 1).value for idx in range(len(header_row))]
        if all(cell in (None, "") for cell in row_values):
            continue
        first_cell_text = _as_text(row_values[0])
        non_empty_count = sum(1 for cell in row_values if cell not in (None, ""))
        if "开始" in first_cell_text and "填写" in first_cell_text:
            continue
        if non_empty_count <= 1 and first_cell_text:
            continue

        row: dict[str, Any] = {}
        for field in ADQ_INPUT_FIELDS:
            raw_value = row_values[header_map[field]]
            required = field not in OPTIONAL_FIELDS
            if field in NUMERIC_FIELDS:
                row[field] = _as_number(raw_value, required=required, field=field, row_idx=row_idx)
            else:
                text = _as_text(raw_value)
                if required and text == "":
                    raise ValueError(f"第 {row_idx} 行缺少必填字段：{field}")
                row[field] = text

        row["直播时长(h)"] = _resolve_duration_hours(row, row_idx)
        data_rows.append(row)

    if not data_rows:
        raise ValueError("输入模板中未找到数据行（请从第 5 行开始填写）")
    return data_rows
