import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from adq_input_loader import load_adq_rows

INPUT_FILE = "data/adq_input_template.xlsx"

wb = openpyxl.Workbook()
rows = load_adq_rows(INPUT_FILE)
primary_session = rows[0]["场次ID"]
session_rows = [row for row in rows if row["场次ID"] == primary_session] or rows
group1 = session_rows[0]
group2 = session_rows[1] if len(session_rows) > 1 else {
    **group1,
    "投流组名": "投放组 2",
    "观看人数": 0.0,
    "花费(元)": 0.0,
    "成交量": 0.0,
    "转化目标成本(元)": 0.0,
    "商品点击": 0.0,
    "提交订单": 0.0,
    "支付成功": 0.0,
    "互动人数": 0.0,
}

# 样式定义
title_fill   = PatternFill("solid", fgColor="1A3A6B")
header_fill  = PatternFill("solid", fgColor="2F4F8F")
green_light  = PatternFill("solid", fgColor="E2EFDA")
yellow_fill  = PatternFill("solid", fgColor="FFF2CC")
gray_fill    = PatternFill("solid", fgColor="F2F2F2")
input_fill   = PatternFill("solid", fgColor="FFFDE7")
orange_light = PatternFill("solid", fgColor="FAD7A0")
blue_light   = PatternFill("solid", fgColor="AED6F1")
pink_fill    = PatternFill("solid", fgColor="FFE0E0")

white_bold  = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def st(ws, cell, val=None, fill=None, font=None, align=center, border=thin, fmt=None):
    c = ws[cell]
    if val is not None:
        c.value = val
    c.fill = fill if fill else PatternFill()
    c.font = font if font else Font(size=10, color="1F2D3D")
    c.alignment = align
    c.border = border
    if fmt:
        c.number_format = fmt
    return c

# Sheet 1: ADQ原始数据录入
ws1 = wb.active
ws1.title = "ADQ原始数据录入"

ws1.merge_cells("A1:P1")
st(ws1, "A1", "ADQ 投放数据录入表", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:P2")
st(ws1, "A2", "说明：已根据截图自动填入左右两组投流数据，绿色单元格为自动计算结果",
   fill=yellow_fill, font=Font(bold=True, color="7B3F00", size=10), align=left)
ws1.row_dimensions[2].height = 20

ws1.merge_cells("A3:P3")
st(ws1, "A3", "基本信息", fill=header_fill, font=white_bold, align=left)
ws1.row_dimensions[3].height = 22

for col, hdr in zip(["A","B","C","D","E"], ["日期","主播","开播时间","结束时间","直播时长(h)"]):
    st(ws1, f"{col}4", hdr, fill=header_fill, font=white_bold)

st(ws1, "A5", group1["日期"], fill=input_fill)
st(ws1, "B5", group1["主播"], fill=input_fill)
st(ws1, "C5", group1["开播时间"], fill=input_fill)
st(ws1, "D5", group1["结束时间"], fill=input_fill)
ws1["C5"].number_format = "@"
ws1["D5"].number_format = "@"
ws1["E5"].value = group1["直播时长(h)"]
ws1["E5"].number_format = "0.0"
ws1["E5"].fill = green_light
ws1["E5"].font = Font(size=10, color="375623")
ws1["E5"].alignment = center
ws1["E5"].border = thin

ws1.merge_cells("A7:H7")
st(ws1, "A7", f"{group1['投流组名']}（组1）", fill=PatternFill("solid", fgColor="E67E22"), font=white_bold)
ws1.merge_cells("I7:P7")
st(ws1, "I7", f"{group2['投流组名']}（组2）", fill=PatternFill("solid", fgColor="2471A3"), font=white_bold)

fields = ["观看人数", "花费(元)", "成交量", "转化目标成本(元)", "商品点击", "提交订单", "支付成功", "互动人数"]
g1_cols = ["A","B","C","D","E","F","G","H"]
g2_cols = ["I","J","K","L","M","N","O","P"]

for col, field in zip(g1_cols, fields):
    st(ws1, f"{col}8", field, fill=orange_light, font=Font(bold=True, size=9, color="7B3F00"))
for col, field in zip(g2_cols, fields):
    st(ws1, f"{col}8", field, fill=blue_light, font=Font(bold=True, size=9, color="1A5276"))

# 根据输入模板自动填入的数据
left_values = [
    group1["观看人数"],
    group1["花费(元)"],
    group1["成交量"],
    group1["转化目标成本(元)"],
    group1["商品点击"],
    group1["提交订单"],
    group1["支付成功"],
    group1["互动人数"],
]
right_values = [
    group2["观看人数"],
    group2["花费(元)"],
    group2["成交量"],
    group2["转化目标成本(元)"],
    group2["商品点击"],
    group2["提交订单"],
    group2["支付成功"],
    group2["互动人数"],
]

for col, val in zip(g1_cols, left_values):
    st(ws1, f"{col}9", val, fill=input_fill)
for col, val in zip(g2_cols, right_values):
    st(ws1, f"{col}9", val, fill=input_fill)

ws1.merge_cells("A10:P10")
st(ws1, "A10", f"输入来源：{INPUT_FILE}；互动人数按截图中的评论人数填写。",
   fill=gray_fill, font=Font(italic=True, size=9, color="666666"), align=left)

for col in "ABCDEFGHIJKLMNOP":
    ws1.column_dimensions[col].width = 14

# Sheet 2: 汇总统计表
ws2 = wb.create_sheet("汇总统计表")
ws2.merge_cells("A1:Q1")
st(ws2, "A1", "直播 ADQ 投放汇总统计表（已填数据版）", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))

summary_headers = [
    "日期","主播","时间","直播时长(h)",
    "消耗(元)","成交量","成本(元/单)",
    "场观(人)","观看成本(元)",
    "商品点击","商品点击率",
    "提交订单","支付成功","未支付",
    "互动人数","转化率","互动转化率"
]
for i, hdr in enumerate(summary_headers):
    col = get_column_letter(i+1)
    st(ws2, f"{col}3", hdr, fill=header_fill, font=white_bold)

ref = "ADQ原始数据录入"
formulas = [
    ("A", f"='{ref}'!A5", None, input_fill),
    ("B", f"='{ref}'!B5", None, input_fill),
    ("C", f"='{ref}'!C5&\"-\"&'{ref}'!D5", None, input_fill),
    ("D", f"='{ref}'!E5", "0.0", input_fill),
    ("E", f"='{ref}'!B9+'{ref}'!J9", "#,##0.00", green_light),
    ("F", f"='{ref}'!C9+'{ref}'!K9", None, green_light),
    ("G", f"=IFERROR(E4/F4,\"\")", "#,##0.00", green_light),
    ("H", f"='{ref}'!A9+'{ref}'!I9", None, green_light),
    ("I", f"=IFERROR(E4/H4,\"\")", "#,##0.00", green_light),
    ("J", f"='{ref}'!E9+'{ref}'!M9", None, green_light),
    ("K", f"=IFERROR(J4/H4,\"\")", "0.00%", green_light),
    ("L", f"='{ref}'!F9+'{ref}'!N9", None, green_light),
    ("M", f"='{ref}'!G9+'{ref}'!O9", None, green_light),
    ("N", f"=IFERROR(L4-M4,\"\")", None, pink_fill),
    ("O", f"='{ref}'!H9+'{ref}'!P9", None, green_light),
    ("P", f"=IFERROR(F4/H4,\"\")", "0.00%", green_light),
    ("Q", f"=IFERROR(F4/O4,\"\")", "0.00%", green_light),
]
for col, formula, fmt, fill in formulas:
    st(ws2, f"{col}4", formula, fill=fill, fmt=fmt)

for i, w in enumerate([12,10,16,12,13,10,13,10,13,10,12,10,10,10,10,10,12]):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# Sheet 3: 固定值展示
ws3 = wb.create_sheet("识别结果")
ws3.merge_cells("A1:H1")
st(ws3, "A1", "本次截图识别结果", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
rows = [
    ["组别", "观看人数", "花费", "成交量", "目标成本", "商品点击", "提交订单", "支付成功/互动人数"],
    [
        group1["投流组名"],
        group1["观看人数"],
        group1["花费(元)"],
        group1["成交量"],
        group1["转化目标成本(元)"],
        group1["商品点击"],
        group1["提交订单"],
        f"{int(group1['支付成功'])} / {int(group1['互动人数'])}",
    ],
    [
        group2["投流组名"],
        group2["观看人数"],
        group2["花费(元)"],
        group2["成交量"],
        group2["转化目标成本(元)"],
        group2["商品点击"],
        group2["提交订单"],
        f"{int(group2['支付成功'])} / {int(group2['互动人数'])}",
    ],
]
for r_idx, row in enumerate(rows, start=2):
    for c_idx, val in enumerate(row, start=1):
        fill = header_fill if r_idx == 2 else input_fill
        font = white_bold if r_idx == 2 else Font(size=10, color="1F2D3D")
        st(ws3, f"{get_column_letter(c_idx)}{r_idx}", val, fill=fill, font=font)
for i in range(1, 9):
    ws3.column_dimensions[get_column_letter(i)].width = 16

wb.save("ADQ直播数据统计_已填数据.xlsx")
print("已生成：ADQ直播数据统计_已填数据.xlsx")
