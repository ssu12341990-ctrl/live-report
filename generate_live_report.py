import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "直播数据统计"

# 样式定义
header_fill  = PatternFill("solid", fgColor="2F4F8F")
yellow_fill  = PatternFill("solid", fgColor="FFF2CC")
green_fill   = PatternFill("solid", fgColor="E2EFDA")
pink_fill    = PatternFill("solid", fgColor="FFE0E0")
gray_fill    = PatternFill("solid", fgColor="F2F2F2")
white_bold   = Font(bold=True, color="FFFFFF", size=11)
dark_bold    = Font(bold=True, color="1F2D3D", size=10)
dark_normal  = Font(color="1F2D3D", size=10)
red_bold     = Font(bold=True, color="C00000", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

# 第1行：标题
ws.merge_cells("A1:R1")
ws["A1"] = "直播间数据统计报告（修正版）"
ws["A1"].font      = Font(bold=True, color="FFFFFF", size=14)
ws["A1"].fill      = PatternFill("solid", fgColor="1A3A6B")
ws["A1"].alignment = center
ws.row_dimensions[1].height = 36

# 第2行：修正说明
ws.merge_cells("A2:R2")
ws["A2"] = "修正说明：① 原[场观roi/总ROI]实为转化率，已重命名；② 新增[成交金额]与[真实ROI(成交额/消耗)]列"
ws["A2"].font      = Font(bold=True, color="C00000", size=10)
ws["A2"].fill      = PatternFill("solid", fgColor="FFF2CC")
ws["A2"].alignment = left
ws.row_dimensions[2].height = 22

# 第3行：列头
headers = [
    ("A", "日期",                   header_fill, white_bold),
    ("B", "时间",                   header_fill, white_bold),
    ("C", "时长(h)",                header_fill, white_bold),
    ("D", "主播",                   header_fill, white_bold),
    ("E", "总场观",                 header_fill, white_bold),
    ("F", "总单数",                 header_fill, white_bold),
    ("G", "消耗(¥)",                header_fill, white_bold),
    ("H", "单位成本(¥)",            header_fill, white_bold),
    ("I", "微信豆场观",             header_fill, white_bold),
    ("J", "微信豆单量",             header_fill, white_bold),
    ("K", "微信豆转化率",           green_fill,  Font(bold=True, color="375623", size=10)),
    ("L", "ADQ场观",                header_fill, white_bold),
    ("M", "ADQ单量",                header_fill, white_bold),
    ("N", "ADQ场观转化率",          green_fill,  Font(bold=True, color="375623", size=10)),
    ("O", "总转化率",               green_fill,  Font(bold=True, color="375623", size=10)),
    ("P", "成交金额(¥)",            yellow_fill, Font(bold=True, color="7B5900", size=10)),
    ("Q", "真实ROI\n(成交额/消耗)", yellow_fill, Font(bold=True, color="7B5900", size=10)),
    ("R", "备注",                   header_fill, white_bold),
]
for col, title, fill, font in headers:
    c = ws[f"{col}3"]
    c.value = title
    c.fill  = fill
    c.font  = font
    c.alignment = center
    c.border    = thin_border
ws.row_dimensions[3].height = 40

# 数据行
rows = [
    {
        "日期": "2024/5/21", "时间": "8:00-10:30", "时长": 2.5, "主播": "王雨琪",
        "总场观": 683, "总单数": 47, "消耗": 10191.95, "单位成本": 216.85,
        "wx豆场观": 0, "wx豆单量": 0, "wx豆转化率": None,
        "adq场观": 452, "adq单量": 42, "adq转化率": 42/452,
        "总转化率": 47/683, "成交金额": 465.30, "真实ROI": 465.30/10191.95,
        "备注": "微信豆未投放",
    },
]

ROW_START = 4
for i, d in enumerate(rows):
    r = ROW_START + i
    ws.row_dimensions[r].height = 22
    mapping = [
        ("A", d["日期"],        dark_normal, None,       None),
        ("B", d["时间"],        dark_normal, None,       None),
        ("C", d["时长"],        dark_normal, None,       None),
        ("D", d["主播"],        dark_bold,   None,       None),
        ("E", d["总场观"],      dark_normal, None,       None),
        ("F", d["总单数"],      dark_normal, None,       None),
        ("G", d["消耗"],        dark_normal, "#,##0.00", None),
        ("H", d["单位成本"],    dark_normal, "#,##0.00", None),
        ("I", d["wx豆场观"],    dark_normal, None,       None),
        ("J", d["wx豆单量"],    dark_normal, None,       None),
        ("K", d["wx豆转化率"],  dark_normal, "0.00%",    green_fill),
        ("L", d["adq场观"],     dark_normal, None,       None),
        ("M", d["adq单量"],     dark_normal, None,       None),
        ("N", d["adq转化率"],   dark_normal, "0.00%",    green_fill),
        ("O", d["总转化率"],    dark_normal, "0.00%",    green_fill),
        ("P", d["成交金额"],    dark_bold,   "#,##0.00", yellow_fill),
        ("Q", d["真实ROI"],     red_bold,    "0.00%",    yellow_fill),
        ("R", d["备注"],        dark_normal, None,       None),
    ]
    for col, val, font, fmt, fill in mapping:
        c = ws[f"{col}{r}"]
        c.font      = font
        c.alignment = center
        c.border    = thin_border
        if fill: c.fill = fill
        if col == "K" and val is None:
            c.value = "#DIV/0!"
            c.font  = Font(color="C00000", bold=True, size=10)
        else:
            c.value = val
            if fmt: c.number_format = fmt

# 列宽
for col, w in zip("ABCDEFGHIJKLMNOPQR",
                  [10,13,8,10,9,9,13,13,10,10,14,9,9,16,12,13,18,18]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A4"

# 字段说明区 —— 纯文字，无Emoji，无特殊符号，只用颜色区分
lr = ROW_START + len(rows) + 2
ws.merge_cells(f"A{lr}:R{lr}")
ws[f"A{lr}"] = "字段说明"
ws[f"A{lr}"].font      = Font(bold=True, size=11, color="FFFFFF")
ws[f"A{lr}"].fill      = PatternFill("solid", fgColor="1A3A6B")
ws[f"A{lr}"].alignment = left
ws.row_dimensions[lr].height = 22

# 注意：说明文字全部用普通符号，不含 ÷ / 特殊数学符号
legends = [
    ("转化率（修正前称roi）", "计算方式：该渠道成交单数 / 该渠道场观人数，反映流量转化效率",         green_fill),
    ("真实ROI",               "计算方式：累计成交金额 / 广告消耗，大于100%才回本，本场仅4.56%亏损",  yellow_fill),
    ("单位成本",               "计算方式：广告消耗 / 总成交单数，即每单获客成本216.85元",             gray_fill),
    ("本场结论",               "成交465.30元 / 消耗10191.95元，建议复盘投放策略与商品定价",           pink_fill),
]
for j, (term, desc, bg) in enumerate(legends):
    row_n = lr + 1 + j
    ws.row_dimensions[row_n].height = 20
    ws.merge_cells(f"A{row_n}:D{row_n}")
    ws.merge_cells(f"E{row_n}:R{row_n}")
    c_term = ws[f"A{row_n}"]
    c_term.value     = term
    c_term.font      = Font(bold=True, size=10, color="1F2D3D")
    c_term.alignment = left
    c_term.fill      = bg
    c_term.border    = thin_border
    c_desc = ws[f"E{row_n}"]
    c_desc.value     = desc
    c_desc.font      = Font(size=10, color="333333")
    c_desc.alignment = left
    c_desc.fill      = bg
    c_desc.border    = thin_border

wb.save("直播数据统计_修正版.xlsx")
print("已生成：直播数据统计_修正版.xlsx")
