from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from adq_input_loader import load_adq_rows

INPUT_FILE = "data/adq_input_template.xlsx"
OUTPUT_FILE = "ADQ高级投流分析_已填数据.xlsx"

rows = load_adq_rows(INPUT_FILE)

wb = openpyxl.Workbook()
title_fill = PatternFill("solid", fgColor="1A3A6B")
header_fill = PatternFill("solid", fgColor="2F4F8F")
input_fill = PatternFill("solid", fgColor="FFFDE7")
green_fill = PatternFill("solid", fgColor="E2EFDA")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
pink_fill = PatternFill("solid", fgColor="FFE0E0")
white_bold = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def st(ws, cell, value=None, fill=None, font=None, align=center, border=thin, fmt=None):
    c = ws[cell]
    if value is not None:
        c.value = value
    c.fill = fill if fill else PatternFill()
    c.font = font if font else Font(size=10, color="1F2D3D")
    c.alignment = align
    c.border = border
    if fmt:
        c.number_format = fmt
    return c


def ratio(a, b):
    return 0 if not b else a / b


# Sheet 1: 投流明细
ws = wb.active
ws.title = "投流明细"
detail_headers = [
    "日期",
    "主播",
    "开播时间",
    "结束时间",
    "直播时长(h)",
    "场次ID",
    "投流组名",
    "投流渠道",
    "创意名称",
    "主要地域",
    "观看人数",
    "花费(元)",
    "成交量",
    "转化目标成本(元)",
    "商品点击",
    "提交订单",
    "支付成功",
    "互动人数",
    "关注人数",
    "支付金额(元)",
    "退款金额(元)",
    "备注",
]
ws.merge_cells("A1:V1")
st(ws, "A1", f"ADQ高级投流分析（数据源：{INPUT_FILE}）", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
for i, header in enumerate(detail_headers, start=1):
    st(ws, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
for row_idx, row in enumerate(rows, start=3):
    for col_idx, header in enumerate(detail_headers, start=1):
        value = row[header]
        fmt = "#,##0.00" if header in {"花费(元)", "转化目标成本(元)", "支付金额(元)", "退款金额(元)"} else None
        st(ws, f"{get_column_letter(col_idx)}{row_idx}", value, fill=input_fill, fmt=fmt)
for i in range(1, 23):
    ws.column_dimensions[get_column_letter(i)].width = 14
ws.freeze_panes = "A3"


# Prepare grouped data
session_groups = defaultdict(list)
date_groups = defaultdict(list)
for row in rows:
    session_key = (row["日期"], row["主播"], row["场次ID"], row["开播时间"], row["结束时间"], row["直播时长(h)"])
    session_groups[session_key].append(row)
    date_groups[(row["日期"], row["主播"])].append(row)


# Sheet 2: 汇总统计
ws2 = wb.create_sheet("汇总统计")
summary_headers = [
    "日期",
    "主播",
    "场次ID",
    "时间段",
    "直播时长(h)",
    "投流组数",
    "总消耗(元)",
    "总成交量",
    "总场观",
    "总商品点击",
    "总提交订单",
    "总支付成功",
    "总互动人数",
    "支付金额(元)",
    "退款金额(元)",
    "净收入(元)",
    "成交率",
    "点击率",
    "支付率",
    "ROAS",
    "ROI",
]
ws2.merge_cells("A1:U1")
st(ws2, "A1", "汇总统计（按场次）", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
for i, header in enumerate(summary_headers, start=1):
    st(ws2, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
for row_idx, (key, group_rows) in enumerate(session_groups.items(), start=3):
    total_cost = sum(r["花费(元)"] for r in group_rows)
    total_deals = sum(r["成交量"] for r in group_rows)
    total_views = sum(r["观看人数"] for r in group_rows)
    total_clicks = sum(r["商品点击"] for r in group_rows)
    total_orders = sum(r["提交订单"] for r in group_rows)
    total_paid = sum(r["支付成功"] for r in group_rows)
    total_interact = sum(r["互动人数"] for r in group_rows)
    paid_amount = sum(r["支付金额(元)"] for r in group_rows)
    refund_amount = sum(r["退款金额(元)"] for r in group_rows)
    net_revenue = paid_amount - refund_amount
    values = [
        key[0],
        key[1],
        key[2],
        f"{key[3]}-{key[4]}",
        key[5],
        len(group_rows),
        total_cost,
        total_deals,
        total_views,
        total_clicks,
        total_orders,
        total_paid,
        total_interact,
        paid_amount,
        refund_amount,
        net_revenue,
        ratio(total_paid, total_views),
        ratio(total_clicks, total_views),
        ratio(total_paid, total_orders),
        ratio(paid_amount, total_cost),
        ratio(net_revenue - total_cost, total_cost),
    ]
    for col_idx, value in enumerate(values, start=1):
        fmt = None
        if col_idx in {7, 14, 15, 16}:
            fmt = "#,##0.00"
        if col_idx in {17, 18, 19, 20, 21}:
            fmt = "0.00%"
        st(ws2, f"{get_column_letter(col_idx)}{row_idx}", value, fill=green_fill, fmt=fmt)
for i in range(1, 22):
    ws2.column_dimensions[get_column_letter(i)].width = 13
ws2.freeze_panes = "A3"


# Sheet 3: 组间对比
ws3 = wb.create_sheet("组间对比")
compare_headers = [
    "日期",
    "主播",
    "场次ID",
    "投流组名",
    "投流渠道",
    "创意名称",
    "花费(元)",
    "观看人数",
    "成交量",
    "商品点击",
    "提交订单",
    "支付成功",
    "互动人数",
    "支付金额(元)",
    "退款金额(元)",
    "净收入(元)",
    "点击率",
    "支付率",
    "成交率",
    "ROAS",
    "ROI",
]
ws3.merge_cells("A1:U1")
st(ws3, "A1", "组间对比（按投流组）", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
for i, header in enumerate(compare_headers, start=1):
    st(ws3, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
for row_idx, row in enumerate(rows, start=3):
    net_revenue = row["支付金额(元)"] - row["退款金额(元)"]
    values = [
        row["日期"],
        row["主播"],
        row["场次ID"],
        row["投流组名"],
        row["投流渠道"],
        row["创意名称"],
        row["花费(元)"],
        row["观看人数"],
        row["成交量"],
        row["商品点击"],
        row["提交订单"],
        row["支付成功"],
        row["互动人数"],
        row["支付金额(元)"],
        row["退款金额(元)"],
        net_revenue,
        ratio(row["商品点击"], row["观看人数"]),
        ratio(row["支付成功"], row["提交订单"]),
        ratio(row["支付成功"], row["观看人数"]),
        ratio(row["支付金额(元)"], row["花费(元)"]),
        ratio(net_revenue - row["花费(元)"], row["花费(元)"]),
    ]
    for col_idx, value in enumerate(values, start=1):
        fmt = None
        if col_idx in {7, 14, 15, 16}:
            fmt = "#,##0.00"
        if col_idx in {17, 18, 19, 20, 21}:
            fmt = "0.00%"
        st(ws3, f"{get_column_letter(col_idx)}{row_idx}", value, fill=input_fill, fmt=fmt)
for i in range(1, 22):
    ws3.column_dimensions[get_column_letter(i)].width = 13
ws3.freeze_panes = "A3"


# Sheet 4: 趋势汇总
ws4 = wb.create_sheet("趋势汇总")
trend_headers = [
    "日期",
    "主播",
    "场次数",
    "投流组数",
    "总消耗(元)",
    "总成交量",
    "总场观",
    "总商品点击",
    "总提交订单",
    "总支付成功",
    "总互动人数",
    "支付金额(元)",
    "退款金额(元)",
    "净收入(元)",
    "成交率",
    "点击率",
    "支付率",
    "ROAS",
    "ROI",
]
ws4.merge_cells("A1:S1")
st(ws4, "A1", "趋势汇总（按日期+主播）", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
for i, header in enumerate(trend_headers, start=1):
    st(ws4, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
for row_idx, ((date, host), date_rows) in enumerate(sorted(date_groups.items()), start=3):
    session_count = len({r["场次ID"] for r in date_rows})
    total_cost = sum(r["花费(元)"] for r in date_rows)
    total_deals = sum(r["成交量"] for r in date_rows)
    total_views = sum(r["观看人数"] for r in date_rows)
    total_clicks = sum(r["商品点击"] for r in date_rows)
    total_orders = sum(r["提交订单"] for r in date_rows)
    total_paid = sum(r["支付成功"] for r in date_rows)
    total_interact = sum(r["互动人数"] for r in date_rows)
    paid_amount = sum(r["支付金额(元)"] for r in date_rows)
    refund_amount = sum(r["退款金额(元)"] for r in date_rows)
    net_revenue = paid_amount - refund_amount
    values = [
        date,
        host,
        session_count,
        len(date_rows),
        total_cost,
        total_deals,
        total_views,
        total_clicks,
        total_orders,
        total_paid,
        total_interact,
        paid_amount,
        refund_amount,
        net_revenue,
        ratio(total_paid, total_views),
        ratio(total_clicks, total_views),
        ratio(total_paid, total_orders),
        ratio(paid_amount, total_cost),
        ratio(net_revenue - total_cost, total_cost),
    ]
    for col_idx, value in enumerate(values, start=1):
        fmt = None
        if col_idx in {5, 12, 13, 14}:
            fmt = "#,##0.00"
        if col_idx in {15, 16, 17, 18, 19}:
            fmt = "0.00%"
        st(ws4, f"{get_column_letter(col_idx)}{row_idx}", value, fill=green_fill, fmt=fmt)
for i in range(1, 20):
    ws4.column_dimensions[get_column_letter(i)].width = 13
ws4.freeze_panes = "A3"


# Sheet 5: 漏斗诊断
ws5 = wb.create_sheet("漏斗诊断")
funnel_headers = [
    "日期",
    "主播",
    "场次ID",
    "投流组名",
    "观看人数",
    "商品点击",
    "提交订单",
    "支付成功",
    "点击率",
    "下单率",
    "支付率",
    "成交率",
    "诊断建议",
]
ws5.merge_cells("A1:M1")
st(ws5, "A1", "漏斗诊断（阈值：点击率5%、支付率70%、成交率8%）", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
for i, header in enumerate(funnel_headers, start=1):
    st(ws5, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
for row_idx, row in enumerate(rows, start=3):
    click_rate = ratio(row["商品点击"], row["观看人数"])
    order_rate = ratio(row["提交订单"], row["商品点击"])
    pay_rate = ratio(row["支付成功"], row["提交订单"])
    deal_rate = ratio(row["支付成功"], row["观看人数"])
    suggestions = []
    if click_rate < 0.05:
        suggestions.append("优化素材/定向")
    if pay_rate < 0.7:
        suggestions.append("检查支付承接")
    if deal_rate < 0.08:
        suggestions.append("复盘话术与商品匹配")
    values = [
        row["日期"],
        row["主播"],
        row["场次ID"],
        row["投流组名"],
        row["观看人数"],
        row["商品点击"],
        row["提交订单"],
        row["支付成功"],
        click_rate,
        order_rate,
        pay_rate,
        deal_rate,
        "；".join(suggestions) if suggestions else "表现稳定",
    ]
    for col_idx, value in enumerate(values, start=1):
        fill = yellow_fill if col_idx == 13 else input_fill
        fmt = "0.00%" if col_idx in {9, 10, 11, 12} else None
        st(ws5, f"{get_column_letter(col_idx)}{row_idx}", value, fill=fill, fmt=fmt, align=left if col_idx == 13 else center)
for i in range(1, 14):
    ws5.column_dimensions[get_column_letter(i)].width = 13
ws5.column_dimensions["M"].width = 30
ws5.freeze_panes = "A3"


# Sheet 6: 流量质量_创意表现
ws6 = wb.create_sheet("流量质量_创意表现")
quality_headers = [
    "日期",
    "主播",
    "场次ID",
    "投流组名",
    "投流渠道",
    "创意名称",
    "主要地域",
    "观看人数",
    "互动人数",
    "关注人数",
    "成交量",
    "互动率",
    "关注率",
    "成交率",
    "备注",
]
ws6.merge_cells("A1:O1")
st(ws6, "A1", "流量质量 / 创意表现", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
for i, header in enumerate(quality_headers, start=1):
    st(ws6, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
for row_idx, row in enumerate(rows, start=3):
    values = [
        row["日期"],
        row["主播"],
        row["场次ID"],
        row["投流组名"],
        row["投流渠道"],
        row["创意名称"],
        row["主要地域"],
        row["观看人数"],
        row["互动人数"],
        row["关注人数"],
        row["成交量"],
        ratio(row["互动人数"], row["观看人数"]),
        ratio(row["关注人数"], row["观看人数"]),
        ratio(row["成交量"], row["观看人数"]),
        row["备注"],
    ]
    for col_idx, value in enumerate(values, start=1):
        fmt = "0.00%" if col_idx in {12, 13, 14} else None
        st(ws6, f"{get_column_letter(col_idx)}{row_idx}", value, fill=input_fill, fmt=fmt, align=left if col_idx in {15} else center)
for i in range(1, 16):
    ws6.column_dimensions[get_column_letter(i)].width = 13
ws6.column_dimensions["O"].width = 24
ws6.freeze_panes = "A3"


# Sheet 7: 指标口径说明
ws7 = wb.create_sheet("指标口径说明")
ws7.merge_cells("A1:D1")
st(ws7, "A1", "指标口径说明", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
meta_headers = ["指标", "定义/计算公式", "字段类型", "备注"]
for i, header in enumerate(meta_headers, start=1):
    st(ws7, f"{get_column_letter(i)}2", header, fill=header_fill, font=white_bold)
meta_rows = [
    ("观看人数", "ADQ截图可见的观看人数", "截图自动识别字段", "必填"),
    ("花费(元)", "ADQ截图消耗金额", "截图自动识别字段", "必填"),
    ("成交量", "投流成交量", "截图自动识别字段", "必填"),
    ("商品点击", "商品点击人数/次数", "截图自动识别字段", "必填"),
    ("提交订单", "提交订单数量", "截图自动识别字段", "必填"),
    ("支付成功", "支付成功数量", "截图自动识别字段", "必填"),
    ("互动人数", "评论/互动人数", "截图自动识别字段", "必填"),
    ("关注人数", "新增关注人数", "人工补录字段", "选填，可留空"),
    ("支付金额(元)", "成交支付金额", "人工补录字段", "选填，可留空"),
    ("退款金额(元)", "退款金额", "人工补录字段", "选填，可留空"),
    ("净收入(元)", "支付金额-退款金额", "自动计算字段", ""),
    ("点击率", "商品点击/观看人数", "自动计算字段", ""),
    ("支付率", "支付成功/提交订单", "自动计算字段", ""),
    ("成交率", "支付成功/观看人数", "自动计算字段", ""),
    ("ROAS", "支付金额/花费", "自动计算字段", ""),
    ("ROI", "(净收入-花费)/花费", "自动计算字段", ""),
]
for row_idx, row in enumerate(meta_rows, start=3):
    for col_idx, value in enumerate(row, start=1):
        fill = pink_fill if value == "人工补录字段" else input_fill
        st(ws7, f"{get_column_letter(col_idx)}{row_idx}", value, fill=fill, align=left if col_idx in {2, 4} else center)
ws7.column_dimensions["A"].width = 16
ws7.column_dimensions["B"].width = 36
ws7.column_dimensions["C"].width = 18
ws7.column_dimensions["D"].width = 24
ws7.freeze_panes = "A3"

wb.save(OUTPUT_FILE)
print(f"已生成：{OUTPUT_FILE}")
