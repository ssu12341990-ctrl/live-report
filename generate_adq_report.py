import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# 样式定义
title_fill   = PatternFill("solid", fgColor="1A3A6B")
header_fill  = PatternFill("solid", fgColor="2F4F8F")
green_light  = PatternFill("solid", fgColor="E2EFDA")
yellow_fill  = PatternFill("solid", fgColor="FFF2CC")
gray_fill    = PatternFill("solid", fgColor="F2F2F2")
input_fill   = PatternFill("solid", fgColor="FFFDE7")
orange_light = PatternFill("solid", fgColor="FAD7A0")
blue_light   = PatternFill("solid", fgColor="AED6F1")
pink_fill    = PatternFill("solid", fgColor="FFE0E0")

white_bold  = Font(bold=True, color="FFFFFF", size=11)
dark_normal = Font(color="1F2D3D", size=10)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def st(ws, cell, val=None, fill=None, font=None, align=center, border=thin, fmt=None):
    c = ws[cell]
    if val is not None: c.value = val
    c.fill      = fill if fill else PatternFill()
    c.font      = font if font else Font(size=10, color="1F2D3D")
    c.alignment = align
    c.border    = border
    if fmt: c.number_format = fmt
    return c

# ================================================
# Sheet 1: ADQ 原始数据录入
# ================================================
ws1 = wb.active
ws1.title = "ADQ原始数据录入"

ws1.merge_cells("A1:P1")
st(ws1, "A1", "ADQ 投放数据录入表", fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:P2")
st(ws1, "A2", "说明：每次直播后将ADQ截图数据填入黄色单元格，绿色单元格自动计算",
   fill=yellow_fill, font=Font(bold=True, color="7B3F00", size=10), align=left)
ws1.row_dimensions[2].height = 20

# 基本信息标题
ws1.merge_cells("A3:P3")
st(ws1, "A3", "基本信息", fill=header_fill, font=white_bold, align=left)
ws1.row_dimensions[3].height = 22

for col, hdr in zip(["A","B","C","D","E"], ["日期","主播","开播时间","结束时间","直播时长(h)"]):
    st(ws1, f"{col}4", hdr, fill=header_fill, font=white_bold)
ws1.row_dimensions[4].height = 28

for col in ["A","B","C","D"]:
    st(ws1, f"{col}5", fill=input_fill)
ws1["C5"].number_format = "HH:MM"
ws1["D5"].number_format = "HH:MM"
# 直播时长自动计算
ws1["E5"].value = "=(D5-C5)*24"
ws1["E5"].number_format = "0.0"
ws1["E5"].fill = green_light
ws1["E5"].font = Font(size=10, color="375623")
ws1["E5"].alignment = center
ws1["E5"].border = thin
ws1.row_dimensions[5].height = 24

# ADQ两组数据
ws1.merge_cells("A7:H7")
st(ws1, "A7", "投放组 1（ADQ左侧截图）",
   fill=PatternFill("solid", fgColor="E67E22"), font=white_bold)
ws1.merge_cells("I7:P7")
st(ws1, "I7", "投放组 2（ADQ右侧截图）",
   fill=PatternFill("solid", fgColor="2471A3"), font=white_bold)
ws1.row_dimensions[7].height = 28

adq_fields = [
    "观看人数", "花费(元)", "成交量", "转化目标成本(元)",
    "商品点击", "提交订单", "支付成功", "互动人数",
]
g1_cols = ["A","B","C","D","E","F","G","H"]
g2_cols = ["I","J","K","L","M","N","O","P"]

for col, field in zip(g1_cols, adq_fields):
    st(ws1, f"{col}8", field, fill=orange_light, font=Font(bold=True, size=9, color="7B3F00"))
    st(ws1, f"{col}9", fill=input_fill)

for col, field in zip(g2_cols, adq_fields):
    st(ws1, f"{col}8", field, fill=blue_light, font=Font(bold=True, size=9, color="1A5276"))
    st(ws1, f"{col}9", fill=input_fill)

ws1.row_dimensions[8].height = 32
ws1.row_dimensions[9].height = 24

ws1.merge_cells("A10:P10")
st(ws1, "A10", "互动人数 = 评论人数 + 喝彩人数（填ADQ截图互动数据合计）",
   fill=gray_fill, font=Font(italic=True, size=9, color="666666"), align=left)
ws1.row_dimensions[10].height = 18

for col in "ABCDEFGHIJKLMNOP":
    ws1.column_dimensions[col].width = 14

ws1.freeze_panes = "A9"

# ================================================
# Sheet 2: 汇总统计表
# ================================================
ws2 = wb.create_sheet("汇总统计表")

ws2.merge_cells("A1:Q1")
st(ws2, "A1", "直播 ADQ 投放汇总统计表",
   fill=title_fill, font=Font(bold=True, color="FFFFFF", size=14))
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:Q2")
st(ws2, "A2", "绿色单元格为自动汇总计算结果，数据来源于【ADQ原始数据录入】，无需手动填写",
   fill=green_light, font=Font(bold=True, color="375623", size=10), align=left)
ws2.row_dimensions[2].height = 20

summary_headers = [
    "日期","主播","时间","直播时长(h)",
    "消耗(元)","成交量","成本(元/单)",
    "场观(人)","观看成本(元)",
    "商品点击","商品点击率",
    "提交订单","支付成功","未支付",
    "互动人数","转化率","互动转化率"
]
for i, hdr in enumerate(summary_headers):
    col = get_column_letter(i+1)
    st(ws2, f"{col}3", hdr, fill=header_fill, font=white_bold)
ws2.row_dimensions[3].height = 36

ref = "ADQ原始数据录入"
auto_cols = {"E","F","G","H","I","J","K","L","M","N","O","P","Q"}
formulas = [
    ("A", f"='{ref}'!A5",                                                            None,          input_fill),
    ("B", f"='{ref}'!B5",                                                            None,          input_fill),
    ("C", f"=TEXT('{ref}'!C5,\"HH:MM\")&\"-\"&TEXT('{ref}'!D5,\"HH:MM\")",          None,          input_fill),
    ("D", f"='{ref}'!E5",                                                            "0.0",         input_fill),
    ("E", f"='{ref}'!B9+'{ref}'!J9",                                                 "#,##0.00",    green_light),
    ("F", f"='{ref}'!C9+'{ref}'!K9",                                                 None,          green_light),
    ("G", f"=IFERROR(E4/F4,\"\")",                                                   "#,##0.00",    green_light),
    ("H", f"='{ref}'!A9+'{ref}'!I9",                                                 None,          green_light),
    ("I", f"=IFERROR(E4/H4,\"\")",                                                   "#,##0.00",    green_light),
    ("J", f"='{ref}'!E9+'{ref}'!M9",                                                 None,          green_light),
    ("K", f"=IFERROR(J4/H4,\"\")",                                                   "0.00%",       green_light),
    ("L", f"='{ref}'!F9+'{ref}'!N9",                                                 None,          green_light),
    ("M", f"='{ref}'!G9+'{ref}'!O9",                                                 None,          green_light),
    ("N", f"=IFERROR(L4-M4,\"\")",                                                   None,          pink_fill),
    ("O", f"='{ref}'!H9+'{ref}'!P9",                                                 None,          green_light),
    ("P", f"=IFERROR(F4/H4,\"\")",                                                   "0.00%",       green_light),
    ("Q", f"=IFERROR(F4/O4,\"\")",                                                   "0.00%",       green_light),
]
for col, formula, fmt, fill in formulas:
    st(ws2, f"{col}4", formula, fill=fill, fmt=fmt)
ws2.row_dimensions[4].height = 24

# 字段说明
ws2.merge_cells("A6:Q6")
st(ws2, "A6", "字段说明", fill=header_fill, font=white_bold, align=left)
ws2.row_dimensions[6].height = 22

legend_data = [
    ("消耗(元)",     "组1花费 + 组2花费",                              gray_fill),
    ("成交量",       "组1成交量 + 组2成交量",                          gray_fill),
    ("成本(元/单)",  "消耗 / 成交量，每单获客成本",                    yellow_fill),
    ("场观(人)",     "组1观看人数 + 组2观看人数",                       gray_fill),
    ("观看成本",     "消耗 / 场观人数",                                yellow_fill),
    ("商品点击率",   "商品点击 / 场观人数",                             yellow_fill),
    ("未支付",       "提交订单 - 支付成功",                             pink_fill),
    ("转化率",       "成交量 / 场观人数，反映整体转化效率",              green_light),
    ("互动转化率",   "成交量 / 互动人数，反映互动后成交效率",            green_light),
]
for j, (term, desc, bg) in enumerate(legend_data):
    r = 7 + j
    ws2.merge_cells(f"A{r}:D{r}")
    ws2.merge_cells(f"E{r}:Q{r}")
    c1 = ws2[f"A{r}"]
    c1.value = term; c1.font = Font(bold=True, size=10); c1.fill = bg
    c1.alignment = left; c1.border = thin
    c2 = ws2[f"E{r}"]
    c2.value = desc; c2.font = Font(size=10, color="444444"); c2.fill = bg
    c2.alignment = left; c2.border = thin
    ws2.row_dimensions[r].height = 20

col_widths2 = [12,10,16,12,13,10,13,10,13,10,12,10,10,10,10,10,12]
for i, w in enumerate(col_widths2):
    ws2.column_dimensions[get_column_letter(i+1)].width = w
ws2.freeze_panes = "A4"

# ================================================
# Sheet 3: 示例数据（占位）
# ================================================
ws3 = wb.create_sheet("示例-20240521")

ws3.merge_cells("A1:Q1")
st(ws3, "A1", "示例：请将本行替换为本次上传输入后的汇总结果",
   fill=PatternFill("solid", fgColor="E67E22"), font=white_bold)
ws3.row_dimensions[1].height = 30

for i, hdr in enumerate(summary_headers):
    col = get_column_letter(i+1)
    st(ws3, f"{col}2", hdr, fill=header_fill, font=white_bold)
ws3.row_dimensions[2].height = 36

tf=0; td=0; tv=0
tc=0; ts=0; tp=0; ti=0

example_vals = [
    ("A", "YYYY/MM/DD",           None,         input_fill),
    ("B", "主播名",               None,         input_fill),
    ("C", "08:00-10:30",           None,         input_fill),
    ("D", 2.5,                    "0.0",        input_fill),
    ("E", tf,                     "#,##0.00",   green_light),
    ("F", td,                     None,         green_light),
    ("G", round(tf/td, 2) if td else "",        "#,##0.00",   green_light),
    ("H", tv,                     None,         green_light),
    ("I", round(tf/tv, 2) if tv else "",        "#,##0.00",   green_light),
    ("J", tc,                     None,         green_light),
    ("K", tc/tv if tv else "",                  "0.00%",      green_light),
    ("L", ts,                     None,         green_light),
    ("M", tp,                     None,         green_light),
    ("N", ts-tp,                  None,         pink_fill),
    ("O", ti,                     None,         green_light),
    ("P", td/tv if tv else "",                  "0.00%",      green_light),
    ("Q", td/ti if ti else "",                  "0.00%",      green_light),
]
for col, val, fmt, fill in example_vals:
    st(ws3, f"{col}3", val, fill=fill, fmt=fmt)
ws3.row_dimensions[3].height = 24

for i, w in enumerate(col_widths2):
    ws3.column_dimensions[get_column_letter(i+1)].width = w
ws3.freeze_panes = "A3"

wb.save("ADQ直播数据统计模板.xlsx")
print("已生成：ADQ直播数据统计模板.xlsx")
